from openpyxl import load_workbook
import os
import math

excel_file = 'sagatave_eksamenam.xlsx'

if not os.path.exists(excel_file):
    print(f"Error: The file '{excel_file}' was not found.")
else:
    wb = load_workbook(excel_file)
    ws = wb['Lapa_0']

    total_price = 0
    count = 0

    for row in range(4, ws.max_row + 1):
        product = ws[f'I{row}'].value
        price = ws[f'K{row}'].value

        if isinstance(product, str) and "LaserJet" in product and isinstance(price, (int, float)):
            total_price += price
            count += 1

    average = math.floor(total_price / count) if count else 0
    print("Answer:", average)
