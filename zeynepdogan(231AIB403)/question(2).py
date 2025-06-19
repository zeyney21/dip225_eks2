from openpyxl import load_workbook
import os

excel_file = 'sagatave_eksamenam.xlsx'

if not os.path.exists(excel_file):
    print(f"Error: The file '{excel_file}' was not found.")
else:
    wb = load_workbook(excel_file)
    ws = wb['Lapa_0']

    count = 0
    for row in range(4, ws.max_row + 1):
        priority = ws[f'H{row}'].value
        delivery_date = ws[f'J{row}'].value

        if priority == "High" and hasattr(delivery_date, "year") and delivery_date.year == 2015:
            count += 1

    print("Answer:", count)
