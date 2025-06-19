from openpyxl import load_workbook
import os
import math

excel_file = 'sagatave_eksamenam.xlsx'

if not os.path.exists(excel_file):
    print(f"Error: The file '{excel_file}' was not found.")
else:
    wb = load_workbook(excel_file)
    ws = wb['Lapa_0']

    total_sum = 0

    for row in range(4, ws.max_row + 1):
        client_type = ws[f'F{row}'].value
        price = ws[f'K{row}'].value
        quantity = ws[f'L{row}'].value
        delivery_fee = ws[f'M{row}'].value

        if (
            client_type == "Korporatīvais" and
            isinstance(quantity, (int, float)) and 40 <= quantity <= 50 and
            isinstance(price, (int, float)) and
            isinstance(delivery_fee, (int, float))
        ):
            total_sum += price * quantity + delivery_fee

    print("Answer:", math.floor(total_sum))
