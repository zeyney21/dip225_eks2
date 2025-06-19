from openpyxl import load_workbook
import os

excel_file = 'sagatave_eksamenam.xlsx'

if not os.path.exists(excel_file):
    print(f"Error: The file '{excel_file}' was not found in the current directory.")
else:
    wb = load_workbook(excel_file)
    ws = wb['Lapa_0']

    count = 0

    for row in range(2, ws.max_row + 1):
        address = ws[f'D{row}'].value
        skaits = ws[f'L{row}'].value

        if address and address.startswith("Ain") and isinstance(skaits, (int, float)) and skaits < 40:
            count += 1

    print("Answer:", count)
