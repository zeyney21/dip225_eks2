from openpyxl import Workbook, load_workbook 
wb=load_workbook('data.xlsx')
ws=wb['Lapa_1']
max_row=ws.max_row
s=[]
for row in range(2,max_row+1):
    a=(ws['a' + str(row)].value)
    s.append(a)
print((s))