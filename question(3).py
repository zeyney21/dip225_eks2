from openpyxl import load_workbook
import os

excel_file = 'sagatave_eksamenam.xlsx'

if not os.path.exists(excel_file):
    print(f"Error: The file '{excel_file}' was not found.")
else:
    wb = load_workbook(excel_file)
    ws = wb['Lapa_0']

    count = 0
    for row in range(4, ws.max_row + 1):
        address = ws[f'D{row}'].value
        city = ws[f'E{row}'].value

        if address and "Adulienas iela" in address and city in ["Valmiera", "Saulkrasti"]:
            count += 1

    print("Answer:", count)
